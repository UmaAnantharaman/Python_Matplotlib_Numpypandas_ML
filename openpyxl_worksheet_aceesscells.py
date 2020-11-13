#Create a workbook using openpyxl library
from openpyxl import Workbook
wb = Workbook()

#Make the current worksheet from the workbook active
ws = wb.active

#Changing the default title of the worksheet
ws.title = "Sample worksheet"

#Iterating through the workbook and printing the sheet names
for sheet in wb:
    print(sheet.title)

#Creating a copy of the worksheet, copy_worksheet usage
source = wb.active
target = wb.copy_worksheet(source)

#Accessing the cells in worksheet by iterating, iter_rows usage
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
